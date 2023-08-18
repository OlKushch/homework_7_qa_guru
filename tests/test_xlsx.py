from openpyxl import load_workbook
from tests.conftest import joined_path
import os
# TODO оформить в тест, добавить ассерты и использовать универсальный путь


def test_workbook_xlsx():
    workbook = load_workbook(os.path.join(joined_path, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)

    assert sheet.cell(row=3, column=2).value == 'Mara'